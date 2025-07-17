import pandas as pd
import os
import glob # This import is not used in the provided functions
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter # This import is not used in the provided functions

subfolder_name = "data" # Define the name of the subfolder for output files

def extract_column_to_csv(excel_file, sheet_name, column_names):
    """
    Extracts specified columns from an Excel file, cleans newlines,
    and saves each as a separate CSV in a 'data' subfolder.
    """
    for column_name in column_names:
        # Create 'data' subfolder if it doesn't exist
        if not os.path.exists(subfolder_name):
            os.makedirs(subfolder_name)
        if not os.path.isfile(os.path.join(subfolder_name, "combined_data.csv")):
            combined_file_path = os.path.join(subfolder_name, "combined_data.csv")
            with open(combined_file_path, 'w') as f:
                pass

        # Determine the output CSV filename based on column name
        if column_name == "Cause":
            output_csv_file = ["Trip_Causes_Uncleaned.csv", "Cause_types.csv"]
        else:
            output_csv_file = f"{column_name}_data.csv"
        
        # Create full path for the output CSV file
        if not isinstance(output_csv_file, list):
            output_csv_file = [output_csv_file]

        for file in output_csv_file:
            output_csv_file_path = os.path.join( subfolder_name, file)

            try:
                # Read the specified Excel sheet into a DataFrame
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                
                # Check if the target column exists in the DataFrame
                if column_name not in df.columns:
                    print(f"Error: Column {column_name} not found in {excel_file}")
                    print(f" Available columns are { df.columns.tolist()}")
                    return # Exit function if column not found

                # Clean newline characters from the selected column
                # Replaces one or more newline/carriage return characters with a single space
                df[column_name] = df[column_name].astype(str).str.replace(r"[\n\r]+", ' ', regex = True)
                
                # Select the column to save
                selected_column = df[column_name]
                if file == "Cause_types.csv":
                    selected_column = selected_column.value_counts().reset_index()

                
                # Save the selected column to CSV
                # index=False prevents writing DataFrame index, header=True includes column name
                selected_column.to_csv(output_csv_file_path, index = False, header=True)

                print(f" Successfully extraced column '{column_name}' from '{excel_file}' \n and saved it to '{file}'") 

            except FileNotFoundError:
                print(f"Error Excel file not found at '{excel_file}")
            except Exception as e:
                print(f"An error occurred: {e}")

def insert_column_to_excel(csv_file_path, excel_file, rows_to_drop):
    """
    Reads a CSV file and inserts its content into the 'Cause' column of an Excel file.
    Assumes 'Cause' column exists and is in "Sheet1".
    """
    try:
        print("inserting clean causes list into excel file...")
        # Load the workbook from the Excel file
        workbook = load_workbook(excel_file)
        # Select the specific sheet by name
        sheet = workbook["Sheet1"] # Assuming the target sheet is always "Sheet1"

        # Delete rows (with NaN in actions and not "No fault found" in causes) from Excel.
        # Add 2 to match Excel's 1-based indexing and skip the header row.
        # Reverse the list to avoid shifting row positions while deleting.
        for row_num in sorted(rows_to_drop, reverse=True):
            sheet.delete_rows(row_num + 2)

        # Find the column number of "Cause" header
        header_row_num = 1 # Assuming headers are in the first row
        target_column = -1 # Initialize with a value indicating not found
        for col_idx, cell in enumerate(sheet[header_row_num]):
            if cell.value == "Cause":
                target_column = col_idx + 1 # Excel columns are 1-indexed
                break
            
        # Read the cleaned CSV file
        # header=None means no header row in CSV
        df_cleaned_cause = pd.read_csv(csv_file_path, header=None)
        # Get the first (and only) column as a list, skipping the first row (assumed header)
        cleaned_values = df_cleaned_cause.iloc[1:, 0].tolist()
            
        # Iterate through the cleaned values and write them to the Excel sheet
        # Starting row for insertion is header_row_num + 1 (below the header)
        for i, cleaned_value in enumerate(cleaned_values):
            target_row = i + header_row_num + 1 # Calculate the target row number
            sheet.cell(column=target_column, row=target_row, value=cleaned_value)

        # Save the modified workbook
        workbook.save(excel_file)

        print(f"Successfully inseted clean Causes into {excel_file}")

    except FileNotFoundError:
        print(f"Error: the excel file {excel_file} was not found")
    except Exception as e:
        print(f"Error: {e}")

def insert_combined_to_excel(csv_file_path, excel_file, sheet_name = "Augmented Data"):
    try:

        df = pd.read_csv(csv_file_path)

        with pd.ExcelWriter(excel_file, engine= 'openpyxl', mode= 'a', if_sheet_exists= 'new') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index = False)
    except FileNotFoundError:
        print(f"Error: the excel file {excel_file} was not found")
    except Exception as e:
        print(f"Error: {e}")